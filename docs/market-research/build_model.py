#!/usr/bin/env python3
"""Build NeuralMind corporate financial model XLSX."""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))
from intel import FINANCIAL, MARKET  # noqa: E402

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="006100")
BOLD_BLACK = Font(bold=True, color="000000")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")

wb = Workbook()
OUTPUT = _HERE / "financial-model.xlsx"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)


def style_header(cell, text, cols=1):
    ws = cell.parent
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    if cols > 1:
        ws.merge_cells(
            start_row=cell.row,
            start_column=cell.column,
            end_row=cell.row,
            end_column=cell.column + cols - 1,
        )
    for c in range(cell.column, cell.column + cols):
        ws.cell(row=cell.row, column=c).fill = HEADER_FILL


def set_input(cell, value, comment_text=""):
    cell.value = value
    cell.font = BLUE
    if comment_text:
        cell.comment = Comment(comment_text, "model")


def set_formula(cell, formula, green=False):
    cell.value = formula
    cell.font = GREEN if green else BLACK


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Inputs
# ═══════════════════════════════════════════════════════════════════════════════
inp = wb.active
inp.title = "Inputs"
inp.sheet_properties.tabColor = "1F4E79"
inp.column_dimensions["A"].width = 32
inp.column_dimensions["B"].width = 20

style_header(inp["A1"], "NEURALMIND — KEY MODEL INPUTS", 4)
inp["A1"].alignment = Alignment(horizontal="center")
style_header(inp["A4"], "MARKET ASSUMPTIONS", 4)

set_input(inp["B6"], MARKET["current_size_2025"], "Source: Business Research Company, 2025")
inp["A6"] = "Total Addressable Market 2025 ($B)"
set_input(inp["B7"], MARKET["projected_2030"], "Source: Mordor Intelligence, 2025")
inp["A7"] = "Projected TAM 2030 ($B)"
set_input(inp["B8"], MARKET["cagr"], "Source: SNS Insider, 2025")
inp["A8"] = "Market CAGR"
inp["B8"].number_format = "0.0%"
set_input(inp["B9"], MARKET["cloud_share_2025"], "Source: Market data, 2025")
inp["A9"] = "Cloud Share of Market 2025"
inp["B9"].number_format = "0.0%"
set_input(inp["B10"], MARKET["enterprise_ai_2024"], "Source: Industry reports, 2024")
inp["A10"] = "Enterprise AI Market 2024 ($B)"
set_input(inp["B11"], MARKET["developer_adoption_2025"], "Source: Industry survey, 2025")
inp["A11"] = "Developer AI Adoption 2025"
inp["B11"].number_format = "0.0%"
set_input(inp["B12"], MARKET["gartner_2028_developer_adoption"], "Source: Gartner forecast, 2024")
inp["A12"] = "Gartner 2028 Developer Adoption"
inp["B12"].number_format = "0.0%"
set_input(inp["B13"], MARKET["code_generation_pct"], "Source: GitHub/Industry data, 2025")
inp["A13"] = "Code Generation Use-Case Share"
inp["B13"].number_format = "0.0%"
set_input(inp["B14"], MARKET["ai_project_cancellation_2027"], "Source: Industry report, 2024")
inp["A14"] = "AI Project Cancellation Rate 2027"
inp["B14"].number_format = "0.0%"

style_header(inp["A16"], "5-YEAR FINANCIAL PROJECTIONS", 4)
for j, lbl in enumerate(["Year", "Revenue ($)", "Cost ($)", "Users (Seats)"], start=1):
    cell = inp.cell(row=18, column=j)
    cell.value = lbl
    cell.font = BOLD_BLACK

for i in range(1, 6):
    row = 18 + i
    yr = f"year_{i}"
    inp.cell(row=row, column=1).value = f"Year {i}"
    inp.cell(row=row, column=1).font = BLACK
    set_input(
        inp.cell(row=row, column=2),
        FINANCIAL[yr]["revenue"],
        f"intel.py: FINANCIAL['{yr}']['revenue']",
    )
    set_input(
        inp.cell(row=row, column=3), FINANCIAL[yr]["cost"], f"intel.py: FINANCIAL['{yr}']['cost']"
    )
    set_input(
        inp.cell(row=row, column=4), FINANCIAL[yr]["users"], f"intel.py: FINANCIAL['{yr}']['users']"
    )

style_header(inp["A25"], "PRICING ASSUMPTIONS", 4)
set_input(inp["B27"], 0, "Free open source")
inp["A27"] = "OSS Price ($/mo)"
set_input(inp["B28"], 25, "intel.py: PRICING['team']")
inp["A28"] = "Team Price ($/user/mo)"
set_input(inp["B29"], 50000, "intel.py: PRICING['enterprise']")
inp["A29"] = "Enterprise Min Price ($/yr)"
set_input(inp["B30"], 75000, "Estimated avg enterprise contract")
inp["A30"] = "Avg Enterprise Deal ($/yr)"
set_input(inp["B31"], 22, "Conservative vs list $25")
inp["A31"] = "Blended Team ARPU ($/user/mo)"
inp["B31"].number_format = "$#,##0"
set_input(inp["B32"], 0.05, "Annual contract increase")
inp["A32"] = "Annual Price Escalator"
inp["B32"].number_format = "0.0%"

style_header(inp["A34"], "GROWTH & MARGIN ASSUMPTIONS", 4)
y2_rev = FINANCIAL["year_2"]["revenue"]
y3_rev = FINANCIAL["year_3"]["revenue"]
inp["A36"] = "Revenue Growth Y2→Y3"
set_input(inp["B36"], (y3_rev - y2_rev) / max(y2_rev, 1), "Derived from intel.py Y2→Y3 revenue")
inp["B36"].number_format = "0.0%"
inp["A37"] = "Gross Margin Target"
set_input(inp["B37"], 0.75, "SaaS benchmark")
inp["B37"].number_format = "0.0%"
inp["A38"] = "R&D % of Revenue"
set_input(inp["B38"], 0.55, "Heavy investment phase")
inp["B38"].number_format = "0.0%"
inp["A39"] = "S&M % of Revenue"
set_input(inp["B39"], 0.30, "Enterprise sales investment")
inp["B39"].number_format = "0.0%"
inp["A40"] = "G&A % of Revenue"
set_input(inp["B40"], 0.15, "Lean team target")
inp["B40"].number_format = "0.0%"
inp["A41"] = "Tax Rate"
set_input(inp["B41"], 0.21, "US federal statutory")
inp["B41"].number_format = "0.0%"
inp["A42"] = "Discount Rate (WACC)"
set_input(inp["B42"], 0.12, "Early-stage venture WACC")
inp["B42"].number_format = "0.0%"

style_header(inp["A44"], "SEGMENT SENSITIVITY", 4)
inp["A46"] = "Engineering Teams — Total"
set_input(inp["B46"], 5000, "Addressable teams")
inp["A47"] = "Engineering Teams — Captured Y5"
set_input(inp["B47"], 0.05, "5% market penetration Y5")
inp["B47"].number_format = "0.0%"
inp["A48"] = "Security-Sensitive — Total"
set_input(inp["B48"], 1200, "Addressable orgs")
inp["A49"] = "Security-Sensitive — Captured Y5"
set_input(inp["B49"], 0.03, "3% market penetration Y5")
inp["B49"].number_format = "0.0%"
inp["A50"] = "Platform/Infra — Total"
set_input(inp["B50"], 3000, "Addressable teams")
inp["A51"] = "Platform/Infra — Captured Y5"
set_input(inp["B51"], 0.04, "4% market penetration Y5")
inp["B51"].number_format = "0.0%"
inp["A52"] = "Business/Management — Total"
set_input(inp["B52"], 8000, "Addressable orgs")
inp["A53"] = "Business/Management — Captured Y5"
set_input(inp["B53"], 0.02, "2% market penetration Y5")
inp["B53"].number_format = "0.0%"

style_header(inp["A55"], "UNIT ECONOMICS", 4)
inp["A57"] = "CAC — Team Segment ($)"
set_input(inp["B57"], 500, "S&M cost / new customers")
inp["A58"] = "CAC — Enterprise Segment ($)"
set_input(inp["B58"], 5000, "Dedicated sales + legal")
inp["A59"] = "Avg Contract Length (months)"
set_input(inp["B59"], 12, "Annual contracts")
inp["A60"] = "Monthly Churn — Team (%)"
set_input(inp["B60"], 0.02, "2% monthly churn")
inp["B60"].number_format = "0.0%"
inp["A61"] = "Monthly Churn — Enterprise (%)"
set_input(inp["B61"], 0.005, "0.5% monthly churn")
inp["B61"].number_format = "0.0%"
inp["A62"] = "Gross Margin (%)"
set_input(inp["B62"], 0.80, "After hosting/infra costs")
inp["B62"].number_format = "0.0%"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: 5-Year Projection (P&L)
# ═══════════════════════════════════════════════════════════════════════════════
proj = wb.create_sheet("5-Year Projection")
proj.sheet_properties.tabColor = "375623"
proj.column_dimensions["A"].width = 30
for col in ["B", "C", "D", "E", "F"]:
    proj.column_dimensions[col].width = 18

style_header(proj["A1"], "5-YEAR P&L PROJECTION", 6)
proj["A1"].alignment = Alignment(horizontal="center")

proj["A3"] = "Line Item"
proj["A3"].font = BOLD_BLACK
for j, yr in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], start=2):
    cell = proj.cell(row=3, column=j)
    cell.value = yr
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

proj["A5"] = "REVENUE DRIVERS"
proj["A5"].font = BOLD_BLACK
proj.merge_cells("A5:F5")
for col in range(1, 7):
    proj.cell(row=5, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

proj["A7"] = "OSS Users (cumulative)"
for j in range(1, 6):
    col = j + 1
    if j == 1:
        set_formula(proj.cell(row=7, column=col), "0")
    elif j == 2:
        set_formula(proj.cell(row=7, column=col), "='5-Year Projection'!B7+250")
    else:
        set_formula(proj.cell(row=7, column=col), f"={get_column_letter(col)}7+250*({j}-1)")

proj["A8"] = "Team Seats"
for j in range(1, 6):
    col = j + 1
    if j == 1:
        set_formula(proj.cell(row=8, column=col), "0")
    else:
        set_formula(proj.cell(row=8, column=col), f"=Inputs!B{18 + j}", green=True)

proj["A9"] = "Enterprise Deals"
for j in range(1, 6):
    col = j + 1
    if j <= 2:
        set_formula(proj.cell(row=9, column=col), "0")
    elif j == 3:
        set_formula(proj.cell(row=9, column=col), "10")
    else:
        prev_col = get_column_letter(col - 1)
        set_formula(proj.cell(row=9, column=col), f"={prev_col}9+ROUND({prev_col}9*0.6,0)")

proj["A11"] = "REVENUE"
proj["A11"].font = BOLD_BLACK

proj["A12"] = "Team Subscription Revenue"
for j in range(1, 6):
    col = j + 1
    if j == 1:
        set_formula(proj.cell(row=12, column=col), "0")
    else:
        set_formula(proj.cell(row=12, column=col), f"={get_column_letter(col)}8*Inputs!$B$31*12")

proj["A13"] = "Enterprise Revenue"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=13, column=col), f"={get_column_letter(col)}9*Inputs!$B$30")

proj["A14"] = "OSS / Grants / Other"
for j in range(1, 6):
    col = j + 1
    if j <= 2:
        set_formula(proj.cell(row=14, column=col), "0")
    else:
        set_formula(proj.cell(row=14, column=col), f"={get_column_letter(col)}8*2")

proj["A15"] = "Total Revenue"
proj["A15"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=15, column=col),
        f"=SUM({get_column_letter(col)}12:{get_column_letter(col)}14)",
    )

proj["A17"] = "Cost of Revenue"
proj["A17"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=17, column=col), f"={get_column_letter(col)}15*(1-Inputs!$B$37)")

proj["A18"] = "Gross Profit"
proj["A18"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=18, column=col), f"={get_column_letter(col)}15-{get_column_letter(col)}17"
    )

proj["A19"] = "Gross Margin %"
proj["A19"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=19, column=col),
        f"=IF({get_column_letter(col)}15=0,0,{get_column_letter(col)}18/{get_column_letter(col)}15)",
    )
    proj.cell(row=19, column=col).number_format = "0.0%"

proj["A21"] = "OPERATING EXPENSES"
proj["A21"].font = BOLD_BLACK
proj.merge_cells("A21:F21")
for col in range(1, 7):
    proj.cell(row=21, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

proj["A23"] = "R&D"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=23, column=col), f"={get_column_letter(col)}15*Inputs!$B$38")
proj["A24"] = "Sales & Marketing"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=24, column=col), f"={get_column_letter(col)}15*Inputs!$B$39")
proj["A25"] = "General & Administrative"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=25, column=col), f"={get_column_letter(col)}15*Inputs!$B$40")
proj["A26"] = "Total OpEx"
proj["A26"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=26, column=col),
        f"=SUM({get_column_letter(col)}23:{get_column_letter(col)}25)",
    )

proj["A28"] = "EBITDA"
proj["A28"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=28, column=col), f"={get_column_letter(col)}18-{get_column_letter(col)}26"
    )
proj["A29"] = "EBITDA Margin %"
proj["A29"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=29, column=col),
        f"=IF({get_column_letter(col)}15=0,0,{get_column_letter(col)}28/{get_column_letter(col)}15)",
    )
    proj.cell(row=29, column=col).number_format = "0.0%"

proj["A31"] = "Depreciation & Amortization"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=31, column=col), f"={get_column_letter(col)}23*0.10")
proj["A32"] = "EBIT"
proj["A32"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=32, column=col), f"={get_column_letter(col)}28-{get_column_letter(col)}31"
    )
proj["A33"] = "Interest Expense"
for j in range(1, 6):
    col = j + 1
    if j <= 2:
        set_formula(proj.cell(row=33, column=col), "0")
    else:
        set_formula(proj.cell(row=33, column=col), "5000")
proj["A34"] = "EBT"
proj["A34"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=34, column=col), f"={get_column_letter(col)}32-{get_column_letter(col)}33"
    )
proj["A35"] = "Tax"
for j in range(1, 6):
    col = j + 1
    set_formula(proj.cell(row=35, column=col), f"=MAX({get_column_letter(col)}34*Inputs!$B$41,0)")
proj["A36"] = "Net Income"
proj["A36"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=36, column=col), f"={get_column_letter(col)}34-{get_column_letter(col)}35"
    )
proj["A37"] = "Net Margin %"
proj["A37"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        proj.cell(row=37, column=col),
        f"=IF({get_column_letter(col)}15=0,0,{get_column_letter(col)}36/{get_column_letter(col)}15)",
    )
    proj.cell(row=37, column=col).number_format = "0.0%"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Pricing Model
# ═══════════════════════════════════════════════════════════════════════════════
pricing = wb.create_sheet("Pricing Model")
pricing.sheet_properties.tabColor = "833C00"
pricing.column_dimensions["A"].width = 28
for col in ["B", "C", "D", "E"]:
    pricing.column_dimensions[col].width = 22

style_header(pricing["A1"], "PRICING MODEL — OSS / TEAM / ENTERPRISE", 5)

pricing["A3"] = "Feature / Metric"
pricing["A3"].font = BOLD_BLACK
for j, tier in enumerate(["OSS", "Team", "Enterprise"], start=2):
    cell = pricing.cell(row=3, column=j)
    cell.value = tier
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

# Build pricing table using explicit cell assignments to avoid formula/quoting issues
pricing_data = [
    # (row_label, oss_value, team_value, ent_value)
    ("Price", "0", "Inputs!$B$28", "Custom"),
    ("Billing", "Free", "Annual", "Annual"),
    ("Users", "Unlimited", "Per-seat", "Custom"),
    ("Graph Intelligence", "✓", "✓", "✓"),
    ("Multi-Language Support", "✓", "✓", "✓"),
    ("RBAC", "—", "✓", "✓"),
    ("SSO", "—", "✓", "✓"),
    ("Audit Trail", "—", "✓", "✓"),
    ("Air-Gap / On-Prem", "Manual", "—", "✓"),
    ("Custom Parsers", "—", "—", "✓"),
    ("Dedicated CSM", "—", "—", "✓"),
    ("SLA", "—", "—", "✓"),
    ("Annual Contract Value", "$0", "Inputs!$B$31*12", "Inputs!$B$30"),
]

for i, (label, oss_val, team_val, ent_val) in enumerate(pricing_data, start=4):
    pricing.cell(row=i, column=1).value = label
    pricing.cell(row=i, column=1).font = (
        BOLD_BLACK if label in ("Price", "Annual Contract Value") else BLACK
    )

    # OSS column (B) - treat as text, blue if numeric
    pricing.cell(row=i, column=2).value = oss_val
    pricing.cell(row=i, column=2).font = BLUE if oss_val.isdigit() else BLACK

    # Team column (C) - green if formula reference, else black
    if team_val.startswith("Inputs!"):
        # Store as formula (openpyxl recognizes strings starting with = as formulas)
        pricing.cell(row=i, column=3).value = f"={team_val}"
        pricing.cell(row=i, column=3).font = GREEN
    else:
        pricing.cell(row=i, column=3).value = team_val
        pricing.cell(row=i, column=3).font = BLACK

    # Enterprise column (D)
    if ent_val.startswith("Inputs!"):
        pricing.cell(row=i, column=4).value = f"={ent_val}"
        pricing.cell(row=i, column=4).font = GREEN
    else:
        pricing.cell(row=i, column=4).value = ent_val
        pricing.cell(row=i, column=4).font = BLACK

# Revenue Mix Projection
style_header(pricing["A19"], "REVENUE MIX PROJECTION", 5)
pricing["A19"].alignment = Alignment(horizontal="center")

pricing["A21"] = "Year"
pricing["A21"].font = BOLD_BLACK
for j in range(1, 6):
    cell = pricing.cell(row=21, column=j + 1)
    cell.value = f"Year {j}"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

pricing["A23"] = "Team ARR"
for j in range(1, 6):
    col = j + 1
    set_formula(
        pricing.cell(row=23, column=col),
        f"='5-Year Projection'!{get_column_letter(col)}12",
        green=True,
    )
pricing["A24"] = "Enterprise ARR"
for j in range(1, 6):
    col = j + 1
    set_formula(
        pricing.cell(row=24, column=col),
        f"='5-Year Projection'!{get_column_letter(col)}13",
        green=True,
    )
pricing["A25"] = "Total ARR"
pricing["A25"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        pricing.cell(row=25, column=col), f"={get_column_letter(col)}23+{get_column_letter(col)}24"
    )
pricing["A27"] = "Team % of Revenue"
pricing["A27"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        pricing.cell(row=27, column=col),
        f"=IF({get_column_letter(col)}25=0,0,{get_column_letter(col)}23/{get_column_letter(col)}25)",
    )
    pricing.cell(row=27, column=col).number_format = "0.0%"
pricing["A28"] = "Enterprise % of Revenue"
pricing["A28"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    set_formula(
        pricing.cell(row=28, column=col),
        f"=IF({get_column_letter(col)}25=0,0,{get_column_letter(col)}24/{get_column_letter(col)}25)",
    )
    pricing.cell(row=28, column=col).number_format = "0.0%"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Segment Sensitivity
# ═══════════════════════════════════════════════════════════════════════════════
sense = wb.create_sheet("Segment Sensitivity")
sense.sheet_properties.tabColor = "7030A0"
sense.column_dimensions["A"].width = 32
for col in ["B", "C", "D", "E", "F"]:
    sense.column_dimensions[col].width = 16

style_header(sense["A1"], "SEGMENT SENSITIVITY ANALYSIS", 6)

# ── Penetration rate sensitivity (5x5) ──
# Layout:
#   Row 3: merged section header (A1 already used, A3 is free)
#   Row 5: column headers (avg prices) B5:F5
#   Rows 6-10: row headers (pen rates) in A, formulas in B:F
style_header(sense["A3"], "Penetration Rate Sensitivity — Team Revenue Y5", 6)

sense.cell(row=5, column=1).value = "Penetration % ↓ / Avg Price →"
sense.cell(row=5, column=1).font = BOLD_BLACK
avg_prices = [15, 20, 25, 30, 35]
for j, p in enumerate(avg_prices, start=2):
    sense.cell(row=5, column=j).value = p
    sense.cell(row=5, column=j).font = BLUE
    sense.cell(row=5, column=j).number_format = "$#,##0"

pen_rates = [0.02, 0.03, 0.05, 0.08, 0.10]
for i, pen in enumerate(pen_rates):
    r = 6 + i
    sense.cell(row=r, column=1).value = pen
    sense.cell(row=r, column=1).font = BLUE
    sense.cell(row=r, column=1).number_format = "0%"
    for j in range(5):
        c = 2 + j
        col_letter = get_column_letter(c)
        # Use cross-sheet references: row header = A{r}, column header = {col_letter}5
        set_formula(sense.cell(row=r, column=c), f"=Inputs!$B$46*A{r}*{col_letter}5*12")
        if pen == 0.05 and j == 2:
            sense.cell(row=r, column=c).fill = LIGHT_BLUE_FILL
            sense.cell(row=r, column=c).font = Font(bold=True, color="1F4E79")

# ── Enterprise deal size sensitivity (5x5) ──
# Layout:
#   Row 12: section header (merged)
#   Row 14: column headers (deals) B14:F14
#   Rows 15-19: row headers (deal sizes) in A, formulas in B:F
style_header(sense["A12"], "Enterprise Deal Size Sensitivity — Revenue Y5", 6)

sense.cell(row=14, column=1).value = "Price ↓ / Deals →"
sense.cell(row=14, column=1).font = BOLD_BLACK
deals = [10, 20, 30, 50, 75]
for j, d in enumerate(deals, start=2):
    sense.cell(row=14, column=j).value = d
    sense.cell(row=14, column=j).font = BLUE

deal_sizes = [30000, 50000, 75000, 100000, 150000]
for i, size in enumerate(deal_sizes):
    r = 15 + i
    sense.cell(row=r, column=1).value = size
    sense.cell(row=r, column=1).font = BLUE
    sense.cell(row=r, column=1).number_format = "$#,##0"
    for j in range(5):
        c = 2 + j
        col_letter = get_column_letter(c)
        set_formula(sense.cell(row=r, column=c), f"=A{r}*{col_letter}14")
        if size == 75000 and j == 2:
            sense.cell(row=r, column=c).fill = LIGHT_BLUE_FILL
            sense.cell(row=r, column=c).font = Font(bold=True, color="1F4E79")

# ── Segment capture summary ──
style_header(sense["A23"], "SEGMENT CAPTURE SUMMARY (Year 5)", 6)

sense.cell(row=25, column=1).value = "Segment"
sense.cell(row=25, column=1).font = BOLD_BLACK
sense.cell(row=25, column=2).value = "Total Addressable"
sense.cell(row=25, column=2).font = BOLD_BLACK
sense.cell(row=25, column=3).value = "Capture Rate"
sense.cell(row=25, column=3).font = BOLD_BLACK
sense.cell(row=25, column=4).value = "Captured Units"
sense.cell(row=25, column=4).font = BOLD_BLACK
sense.cell(row=25, column=5).value = "Avg Revenue/Unit"
sense.cell(row=25, column=5).font = BOLD_BLACK
sense.cell(row=25, column=6).value = "Total Revenue"
sense.cell(row=25, column=6).font = BOLD_BLACK

segments_data = [
    ("Engineering Teams", "B46", "B47", "Inputs!$B$31*12"),
    ("Security-Sensitive", "B48", "B49", "Inputs!$B$30"),
    ("Platform/Infra", "B50", "B51", "Inputs!$B$31*12"),
    ("Business/Management", "B52", "B53", "Inputs!$B$30*0.5"),
]

for i, (name, total_ref, cap_ref, arpu_formula) in enumerate(segments_data, start=26):
    sense.cell(row=i, column=1).value = name
    sense.cell(row=i, column=1).font = BOLD_BLACK
    set_formula(sense.cell(row=i, column=2), f"=Inputs!{total_ref}", green=True)
    set_formula(sense.cell(row=i, column=3), f"=Inputs!{cap_ref}", green=True)
    set_formula(sense.cell(row=i, column=4), f"=B{i}*C{i}")
    set_formula(sense.cell(row=i, column=5), f"={arpu_formula}", green=True)
    set_formula(sense.cell(row=i, column=6), f"=D{i}*E{i}")

sense.cell(row=30, column=1).value = "Total"
sense.cell(row=30, column=1).font = BOLD_BLACK
for c in range(2, 7):
    col_letter = get_column_letter(c)
    set_formula(sense.cell(row=30, column=c), f"=SUM({col_letter}26:{col_letter}29)")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Unit Economics
# ═══════════════════════════════════════════════════════════════════════════════
unit = wb.create_sheet("Unit Economics")
unit.sheet_properties.tabColor = "BF8F00"
unit.column_dimensions["A"].width = 30
for col in ["B", "C"]:
    unit.column_dimensions[col].width = 20

style_header(unit["A1"], "UNIT ECONOMICS", 3)

style_header(unit["A3"], "TEAM TIER", 3)
unit["A5"] = "Average Revenue Per User (ARPU, $/mo)"
unit["B5"].value = "Inputs!$B$31"
unit["B5"].font = GREEN
unit["A6"] = "Monthly Churn Rate"
unit["B6"].value = "Inputs!$B$60"
unit["B6"].font = GREEN
unit["B6"].number_format = "0.0%"
unit["A7"] = "Gross Margin"
unit["B7"].value = "Inputs!$B$62"
unit["B7"].font = GREEN
unit["B7"].number_format = "0.0%"
unit["A8"] = "Customer Lifetime (months)"
set_formula(unit["B8"], "=IF(B6=0,0,1/B6)")
unit["A9"] = "Lifetime Revenue ($)"
set_formula(unit["B9"], "=B5*B8")
unit["A10"] = "LTV (Gross Profit, $)"
set_formula(unit["B10"], "=B9*B7")
unit["A11"] = "CAC ($)"
unit["B11"].value = "Inputs!$B$57"
unit["B11"].font = GREEN
unit["A12"] = "LTV/CAC Ratio"
set_formula(unit["B12"], "=IF(B11=0,0,B10/B11)")
unit["B12"].number_format = "0.0x"
unit["A13"] = "Payback Period (months)"
set_formula(unit["B13"], "=IF(B5=0,0,B11/(B5*B7))")
unit["B13"].number_format = "0.0"

style_header(unit["A16"], "ENTERPRISE TIER", 3)
unit["A18"] = "Average Contract Value ($/yr)"
unit["B18"].value = "Inputs!$B$30"
unit["B18"].font = GREEN
unit["B18"].number_format = "$#,##0"
unit["A19"] = "Monthly Churn Rate"
unit["B19"].value = "Inputs!$B$61"
unit["B19"].font = GREEN
unit["B19"].number_format = "0.0%"
unit["A20"] = "Gross Margin"
unit["B20"].value = "Inputs!$B$62"
unit["B20"].font = GREEN
unit["B20"].number_format = "0.0%"
unit["A21"] = "Customer Lifetime (months)"
set_formula(unit["B21"], "=IF(B19=0,0,1/B19)")
unit["A22"] = "Lifetime Revenue ($)"
set_formula(unit["B22"], "=B18/12*B21")
unit["A23"] = "LTV (Gross Profit, $)"
set_formula(unit["B23"], "=B22*B20")
unit["A24"] = "CAC ($)"
unit["B24"].value = "Inputs!$B$58"
unit["B24"].font = GREEN
unit["A25"] = "LTV/CAC Ratio"
set_formula(unit["B25"], "=IF(B24=0,0,B23/B24)")
unit["B25"].number_format = "0.0x"
unit["A26"] = "Payback Period (months)"
set_formula(unit["B26"], "=IF(B18=0,0,B24/(B18/12*B20))")
unit["B26"].number_format = "0.0"

style_header(unit["A29"], "UNIT ECONOMICS COMPARISON", 3)
unit["A31"] = "Metric"
unit["A31"].font = BOLD_BLACK
unit["B31"] = "Team"
unit["B31"].font = BOLD_BLACK
unit["C31"] = "Enterprise"
unit["C31"].font = BOLD_BLACK

comparisons = [
    ("ARPU / ACV", "B5", "B18"),
    ("Churn (%)", "B6", "B19"),
    ("Gross Margin (%)", "B7", "B20"),
    ("Lifetime (mo)", "B8", "B21"),
    ("LTV ($)", "B10", "B23"),
    ("CAC ($)", "B11", "B24"),
    ("LTV/CAC (x)", "B12", "B25"),
    ("Payback (mo)", "B13", "B26"),
]

for i, (label, team_ref, ent_ref) in enumerate(comparisons, start=32):
    unit.cell(row=i, column=1).value = label
    unit.cell(row=i, column=1).font = BOLD_BLACK
    set_formula(unit.cell(row=i, column=2), f"={team_ref}")
    set_formula(unit.cell(row=i, column=3), f"={ent_ref}")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Checks
# ═══════════════════════════════════════════════════════════════════════════════
chk = wb.create_sheet("Checks")
chk.sheet_properties.tabColor = "C00000"
chk.column_dimensions["A"].width = 36
chk.column_dimensions["B"].width = 22
chk.column_dimensions["C"].width = 12

style_header(chk["A1"], "BALANCE CHECKS & RECONCILIATION", 5)

# ── Balance Sheet Tie ──
style_header(chk["A3"], "BALANCE SHEET BALANCE (Assets = Liabilities + Equity)", 5)

chk["A5"] = "Year"
chk["A5"].font = BOLD_BLACK
for j in range(1, 6):
    cell = chk.cell(row=5, column=j + 1)
    cell.value = f"Year {j}"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

chk["A7"] = "Cash & Equivalents"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    if j == 1:
        set_formula(chk.cell(row=7, column=col), f"='5-Year Projection'!{col_letter}36+100000")
    else:
        prev_col = get_column_letter(col - 1)
        set_formula(chk.cell(row=7, column=col), f"={prev_col}7+'5-Year Projection'!{col_letter}36")

chk["A8"] = "Accounts Receivable"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=8, column=col), f"='5-Year Projection'!{col_letter}15*0.15")

chk["A9"] = "Total Assets"
chk["A9"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=9, column=col), f"={col_letter}7+{col_letter}8")

chk["A11"] = "Accounts Payable"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=11, column=col), f"='5-Year Projection'!{col_letter}26*0.10")

chk["A12"] = "Deferred Revenue"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=12, column=col), f"='5-Year Projection'!{col_letter}15*0.20")

chk["A13"] = "Total Liabilities"
chk["A13"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=13, column=col), f"={col_letter}11+{col_letter}12")

chk["A15"] = "Retained Earnings"
chk["A15"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    if j == 1:
        set_formula(chk.cell(row=15, column=col), f"='5-Year Projection'!{col_letter}36")
    else:
        prev_col = get_column_letter(col - 1)
        set_formula(
            chk.cell(row=15, column=col), f"={prev_col}15+'5-Year Projection'!{col_letter}36"
        )

chk["A16"] = "Total Equity"
chk["A16"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=16, column=col), f"={col_letter}15")

chk["A18"] = "Total Liabilities + Equity"
chk["A18"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=18, column=col), f"={col_letter}13+{col_letter}16")

chk["A20"] = "BS Balance Check (Assets - L&E = 0)"
chk["A20"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=20, column=col), f"={col_letter}9-{col_letter}18")

chk["A21"] = "BS Balanced?"
chk["A21"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=21, column=col), f"=ABS({col_letter}20)<0.01")

# ── Cash Flow Reconciliation ──
style_header(chk["A24"], "CASH FLOW RECONCILIATION", 5)

chk.cell(row=26, column=1).value = "Year"
chk.cell(row=26, column=1).font = BOLD_BLACK
for j in range(1, 6):
    cell = chk.cell(row=26, column=j + 1)
    cell.value = f"Year {j}"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

chk["A28"] = "Net Income"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=28, column=col), f"='5-Year Projection'!{col_letter}36", green=True)
chk["A29"] = "Depreciation Add-Back"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=29, column=col), f"='5-Year Projection'!{col_letter}31", green=True)
chk["A30"] = "Change in Working Capital"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=30, column=col), f"=-'5-Year Projection'!{col_letter}15*0.05")
chk["A31"] = "Operating Cash Flow"
chk["A31"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=31, column=col), f"={col_letter}28+{col_letter}29+{col_letter}30")

chk["A33"] = "Cash Flow = Cash Change?"
chk["A33"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    if j == 1:
        set_formula(chk.cell(row=33, column=col), f"=ABS({col_letter}31-({col_letter}7-100000))<1")
    else:
        prev_col = get_column_letter(col - 1)
        set_formula(
            chk.cell(row=33, column=col), f"=ABS({col_letter}31-({col_letter}7-{prev_col}7))<1"
        )

# ── Revenue Cross-Check ──
style_header(chk["A36"], "REVENUE CROSS-CHECK: P&L vs PRICING MODEL", 5)

chk.cell(row=38, column=1).value = "Year"
chk.cell(row=38, column=1).font = BOLD_BLACK
for j in range(1, 6):
    cell = chk.cell(row=38, column=j + 1)
    cell.value = f"Year {j}"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

chk["A40"] = "Revenue from P&L"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=40, column=col), f"='5-Year Projection'!{col_letter}15", green=True)
chk["A41"] = "Revenue from Pricing Sheet"
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=41, column=col), f"='Pricing Model'!{col_letter}25", green=True)
chk["A42"] = "Difference"
chk["A42"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=42, column=col), f"={col_letter}40-{col_letter}41")
chk["A43"] = "Tied?"
chk["A43"].font = BOLD_BLACK
for j in range(1, 6):
    col = j + 1
    col_letter = get_column_letter(col)
    set_formula(chk.cell(row=43, column=col), f"=ABS({col_letter}42)<0.01")

# ── Rogue Hardcode Check ──
style_header(chk["A46"], "ROGUE HARDCODE CHECK", 5)

chk["A48"] = "All P&L cells are formulas?"
chk["A48"].font = BOLD_BLACK
chk["B48"] = "Manual review required"
chk["B48"].font = BLUE
chk["B48"].comment = Comment(
    "Verify no hardcoded numbers exist in rows 12-37 of '5-Year Projection'", "model"
)

chk["A49"] = "Named ranges defined?"
chk["A49"].font = BOLD_BLACK
chk["B49"] = "See Defined Names"
chk["B49"].font = GREEN

chk["A50"] = "Inputs all blue (#0000FF)?"
chk["A50"].font = BOLD_BLACK
chk["B50"] = "Manual review required"
chk["B50"].font = BLUE
chk["B50"].comment = Comment("Scan Inputs!B6:B62 — all must be blue font", "model")

# ── Named Range Targets ──
style_header(chk["A52"], "KEY FIGURES (Named Range Targets)", 5)

key_figures = [
    ("Market CAGR", "Inputs!$B$8"),
    ("Discount Rate", "Inputs!$B$42"),
    ("Revenue Y5", "'5-Year Projection'!F$15"),
    ("EBITDA Y5", "'5-Year Projection'!F$28"),
    ("Net Income Y5", "'5-Year Projection'!F$36"),
    ("Team ARPU", "Inputs!$B$31"),
    ("Enterprise ACV", "Inputs!$B$30"),
    ("Gross Margin", "Inputs!$B$37"),
    ("Team LTV", "Unit Economics!B$10"),
    ("Enterprise LTV", "Unit Economics!B$23"),
]

for i, (name, ref) in enumerate(key_figures, start=54):
    chk.cell(row=i, column=1).value = name
    chk.cell(row=i, column=1).font = BOLD_BLACK
    chk.cell(row=i, column=2).value = ref
    chk.cell(row=i, column=2).font = GREEN

# Define named ranges
named_ranges = [
    ("Market_CAGR", "Inputs!$B$8"),
    ("Discount_Rate", "Inputs!$B$42"),
    ("Revenue_Y5", "'5-Year Projection'!F$15"),
    ("EBITDA_Y5", "'5-Year Projection'!F$28"),
    ("Net_Income_Y5", "'5-Year Projection'!F$36"),
    ("Team_ARPU", "Inputs!$B$31"),
    ("Enterprise_ACV", "Inputs!$B$30"),
    ("Gross_Margin", "Inputs!$B$37"),
    ("Team_LTV", "'Unit Economics'!B$10"),
    ("Enterprise_LTV", "'Unit Economics'!B$23"),
]

for name, ref in named_ranges:
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names[name] = dn

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"Named ranges: {list(wb.defined_names.keys())}")
